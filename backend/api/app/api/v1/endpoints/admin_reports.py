# api/app/api/v1/endpoints/admin_reports.py
from uuid import UUID
from io import BytesIO
from openpyxl import Workbook
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, Path, Response
from fastapi.responses import StreamingResponse
from starlette.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.db.session import get_db
from app.api.deps.admin import require_admin
from app.models.encuesta import Survey

from app.schemas.admin_reports import (
    SectionScore, SummaryOut, QuestionRowOut,
    TeacherRowOut, TeacherQBreakdown, CommentRow,
    AttemptAnswerRow, QuestionTeacherDetailOut,
    QuestionGlobalOut, QuestionByTeacherRow, QuestionDetailOut,
    CommentListItem, CommentListOut, TeacherDetailOut, ProgressDay, 
    ProgressDailyOut, SectionSummaryRow, TopBottomQuestionRow, 
    TopBottomQuestionsOut, TeacherMatrixOut, TeacherMatrixRow, 
    TeacherFilterItem, SectionFilterItem, QuestionFilterItem, DateRange, FiltersOut,
    TeacherSectionsOut, TeacherSectionScore, StudentHeatmapOut, StudentHeatmapRow
)

import csv, io, json
router = APIRouter(prefix="/reports", tags=["admin-reports"])

def _ensure_survey(db: Session, survey_id: UUID):
    s = db.query(Survey).filter(Survey.id == survey_id).first()
    if not s:
        raise HTTPException(404, "Encuesta no encontrada")
    return s

# 1) SUMMARY
@router.get("/summary", response_model=SummaryOut)
def summary(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    totals = db.execute(text("""
        SELECT
          SUM(CASE WHEN a.estado = 'enviado'     THEN 1 ELSE 0 END) AS enviados,
          SUM(CASE WHEN a.estado = 'en_progreso' THEN 1 ELSE 0 END) AS en_progreso
        FROM public.attempts a
        WHERE a.survey_id = :sid
    """), {"sid": str(survey_id)}).mappings().first() or {}

    enviados = int(totals.get("enviados") or 0)
    en_progreso = int(totals.get("en_progreso") or 0)

    pend_row = db.execute(text("""
        WITH assigned AS (
          SELECT COUNT(*) AS n FROM public.survey_teacher_assignments WHERE survey_id = :sid
        ),
        responded AS (
          SELECT COUNT(DISTINCT a.teacher_id) AS n
          FROM public.attempts a
          WHERE a.survey_id = :sid AND a.estado = 'enviado'
        )
        SELECT
          (SELECT n FROM assigned) AS total_docentes,
          (SELECT n FROM responded) AS responded_docentes,
          GREATEST((SELECT n FROM assigned) - (SELECT n FROM responded), 0) AS pendientes
    """), {"sid": str(survey_id)}).mappings().first() or {}

    total_docentes = int(pend_row.get("total_docentes") or 0)
    responded_docentes = int(pend_row.get("responded_docentes") or 0)
    pendientes = int(pend_row.get("pendientes") or 0)

    global_row = db.execute(text("""
        SELECT AVG(r.valor_likert::numeric) AS score
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
    """), {"sid": str(survey_id)}).mappings().first() or {}

    score_global = float(global_row.get("score")) if global_row.get("score") is not None else None
    completion_rate = float(responded_docentes) / max(total_docentes, 1)

    sec_rows = db.execute(text("""
        SELECT s.id AS section_id, s.titulo, AVG(r.valor_likert::numeric) AS score
        FROM public.responses r
        JOIN public.attempts a   ON a.id = r.attempt_id
        JOIN public.questions q  ON q.id = r.question_id
        JOIN public.survey_sections s ON s.id = q.section_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
        GROUP BY s.id, s.titulo
        ORDER BY s.titulo
    """), {"sid": str(survey_id)}).mappings().all()

    secciones = [
        SectionScore(section_id=r["section_id"], titulo=r["titulo"],
                     score=float(r["score"]) if r["score"] is not None else None)
        for r in sec_rows
    ]

    return SummaryOut(
        enviados=enviados,
        en_progreso=en_progreso,
        pendientes=pendientes,
        completion_rate=completion_rate,
        score_global=score_global,
        secciones=secciones,
    )

# 2) PREGUNTAS (distribución 1..5)
@router.get("/questions", response_model=List[QuestionRowOut])
def questions_summary(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    rows = db.execute(text("""
        SELECT
          q.id AS question_id,
          q.codigo,
          q.enunciado,
          q.orden,
          s.titulo AS section,
          COUNT(r.valor_likert) AS n,
          AVG(r.valor_likert::numeric) AS mean,
          PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY r.valor_likert) AS median,
          STDDEV_POP(r.valor_likert::numeric) AS stddev,
          MIN(r.valor_likert) AS min,
          MAX(r.valor_likert) AS max,
          SUM(CASE WHEN r.valor_likert = 1 THEN 1 ELSE 0 END) AS c1,
          SUM(CASE WHEN r.valor_likert = 2 THEN 1 ELSE 0 END) AS c2,
          SUM(CASE WHEN r.valor_likert = 3 THEN 1 ELSE 0 END) AS c3,
          SUM(CASE WHEN r.valor_likert = 4 THEN 1 ELSE 0 END) AS c4,
          SUM(CASE WHEN r.valor_likert = 5 THEN 1 ELSE 0 END) AS c5
        FROM public.responses r
        JOIN public.attempts a   ON a.id = r.attempt_id
        JOIN public.questions q  ON q.id = r.question_id
        JOIN public.survey_sections s ON s.id = q.section_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
        GROUP BY q.id, q.codigo, q.enunciado, q.orden, s.titulo
        ORDER BY q.orden
    """), {"sid": str(survey_id)}).mappings().all()

    return [QuestionRowOut(**dict(r)) for r in rows]

@router.get("/questions/top-bottom", response_model=TopBottomQuestionsOut)
def questions_top_bottom(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    limit: int = Query(5, ge=1, le=50, description="Tamaño de los listados top y bottom"),
    min_n: int = Query(10, ge=0, description="Mínimo de respuestas por pregunta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Retorna Top/Bottom preguntas por promedio (solo tipo 'likert').
    Filtra por un número mínimo de respuestas (min_n) para evitar sesgos.
    """
    _ensure_survey(db, survey_id)

    base_sql = """
        WITH base AS (
          SELECT
            q.id   AS question_id,
            q.codigo,
            q.enunciado,
            s.titulo AS section,
            COUNT(r.valor_likert)              AS n,
            AVG(r.valor_likert::numeric)       AS avg
          FROM public.attempts a
          JOIN public.responses r ON r.attempt_id = a.id
          JOIN public.questions q ON q.id = r.question_id
          JOIN public.survey_sections s ON s.id = q.section_id
          WHERE a.survey_id = :sid
            AND a.estado = 'enviado'
            AND r.valor_likert IS NOT NULL
            AND q.tipo = 'likert'
          GROUP BY q.id, q.codigo, q.enunciado, s.titulo
        )
    """

    top_rows = db.execute(text(base_sql + """
        SELECT question_id, codigo, enunciado, section, n, avg
        FROM base
        WHERE n >= :min_n
        ORDER BY avg DESC NULLS LAST, codigo
        LIMIT :limit
    """), {"sid": str(survey_id), "min_n": min_n, "limit": limit}).mappings().all()

    bottom_rows = db.execute(text(base_sql + """
        SELECT question_id, codigo, enunciado, section, n, avg
        FROM base
        WHERE n >= :min_n
        ORDER BY avg ASC NULLS LAST, codigo
        LIMIT :limit
    """), {"sid": str(survey_id), "min_n": min_n, "limit": limit}).mappings().all()

    return TopBottomQuestionsOut(
        top=[TopBottomQuestionRow(**dict(r)) for r in top_rows],
        bottom=[TopBottomQuestionRow(**dict(r)) for r in bottom_rows],
    )

# 3) QUESTION – detail (por docente)
@router.get("/questions/{question_id}", response_model=QuestionDetailOut)
def question_detail(
    question_id: UUID = Path(..., description="ID de la pregunta"),
    survey_id: UUID = Query(..., description="ID de encuesta"),
    by: str = Query("teacher", description="Por ahora solo 'teacher'"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    if by and by != "teacher":
        raise HTTPException(status_code=400, detail="Parámetro 'by' solo soporta 'teacher'")

    _ensure_survey(db, survey_id)

    qmeta = db.execute(text("""
        SELECT q.id AS question_id, q.codigo, q.enunciado, s.titulo AS section
        FROM public.questions q
        JOIN public.survey_sections s ON s.id = q.section_id
        WHERE q.id = :qid AND q.survey_id = :sid
    """), {"qid": str(question_id), "sid": str(survey_id)}).mappings().first()

    if not qmeta:
        raise HTTPException(status_code=404, detail="Pregunta no encontrada en esta encuesta")

    g = db.execute(text("""
        SELECT
          COUNT(r.valor_likert) AS n,
          AVG(r.valor_likert::numeric) AS avg,
          SUM(CASE WHEN r.valor_likert = 1 THEN 1 ELSE 0 END) AS c1,
          SUM(CASE WHEN r.valor_likert = 2 THEN 1 ELSE 0 END) AS c2,
          SUM(CASE WHEN r.valor_likert = 3 THEN 1 ELSE 0 END) AS c3,
          SUM(CASE WHEN r.valor_likert = 4 THEN 1 ELSE 0 END) AS c4,
          SUM(CASE WHEN r.valor_likert = 5 THEN 1 ELSE 0 END) AS c5
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.question_id = :qid
          AND r.valor_likert IS NOT NULL
    """), {"sid": str(survey_id), "qid": str(question_id)}).mappings().first() or {}

    global_out = QuestionGlobalOut(
        n = int(g.get("n") or 0),
        avg = float(g["avg"]) if g.get("avg") is not None else None,
        dist = {
            "1": int(g.get("c1") or 0),
            "2": int(g.get("c2") or 0),
            "3": int(g.get("c3") or 0),
            "4": int(g.get("c4") or 0),
            "5": int(g.get("c5") or 0),
        },
    )

    rows = db.execute(text("""
        SELECT a.teacher_id, t.nombre AS teacher_nombre,
               COUNT(r.valor_likert) AS n,
               AVG(r.valor_likert::float) AS avg
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        JOIN public.teachers t ON t.id = a.teacher_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.question_id = :qid
          AND r.valor_likert IS NOT NULL
        GROUP BY a.teacher_id, t.nombre
        ORDER BY avg DESC NULLS LAST, t.nombre
    """), {"sid": str(survey_id), "qid": str(question_id)}).mappings().all()

    by_teacher = [
        QuestionByTeacherRow(
            teacher_id = r["teacher_id"],
            teacher_nombre = r["teacher_nombre"],
            n = int(r["n"] or 0),
            avg = float(r["avg"]) if r.get("avg") is not None else None,
        )
        for r in rows
    ]

    return QuestionDetailOut(**{
        "question_id": qmeta["question_id"],
        "codigo": qmeta["codigo"],
        "enunciado": qmeta["enunciado"],
        "section": qmeta["section"],
        "global": global_out,     # alias
        "by_teacher": by_teacher,
    })

# 3b) Detalle pregunta por docente
@router.get("/questions/{question_id}/teacher/{teacher_id}", response_model=QuestionTeacherDetailOut)
def question_by_teacher(
    question_id: UUID = Path(...),
    teacher_id: UUID = Path(...),
    survey_id: UUID = Query(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    qmeta = db.execute(text("""
        SELECT q.id AS question_id, q.codigo, q.enunciado, s.titulo AS section
        FROM public.questions q
        JOIN public.survey_sections s ON s.id = q.section_id
        WHERE q.id = :qid
    """), {"qid": str(question_id)}).mappings().first()
    if not qmeta:
        raise HTTPException(status_code=404, detail="Pregunta no encontrada")

    trow = db.execute(text("""
        SELECT t.id, t.nombre
        FROM public.teachers t
        JOIN public.survey_teacher_assignments sta ON sta.teacher_id = t.id
        WHERE sta.survey_id = :sid AND t.id = :tid
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().first()
    if not trow:
        raise HTTPException(status_code=404, detail="Docente no asignado a esta encuesta")

    agg = db.execute(text("""
        SELECT
          COUNT(r.valor_likert) AS n,
          AVG(r.valor_likert::numeric) AS avg,
          SUM(CASE WHEN r.valor_likert = 1 THEN 1 ELSE 0 END) AS c1,
          SUM(CASE WHEN r.valor_likert = 2 THEN 1 ELSE 0 END) AS c2,
          SUM(CASE WHEN r.valor_likert = 3 THEN 1 ELSE 0 END) AS c3,
          SUM(CASE WHEN r.valor_likert = 4 THEN 1 ELSE 0 END) AS c4,
          SUM(CASE WHEN r.valor_likert = 5 THEN 1 ELSE 0 END) AS c5
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        WHERE a.survey_id = :sid
          AND a.estado   = 'enviado'
          AND a.teacher_id = :tid
          AND r.question_id = :qid
          AND r.valor_likert IS NOT NULL
    """), {"sid": str(survey_id), "tid": str(teacher_id), "qid": str(question_id)}).mappings().first() or {}

    n  = int(agg.get("n") or 0)
    avg = float(agg.get("avg")) if agg.get("avg") is not None else None
    dist = {
        "1": int(agg.get("c1") or 0),
        "2": int(agg.get("c2") or 0),
        "3": int(agg.get("c3") or 0),
        "4": int(agg.get("c4") or 0),
        "5": int(agg.get("c5") or 0),
    }

    attempts = db.execute(text("""
        SELECT
          a.id AS attempt_id,
          to_char(COALESCE(r.created_at, a.actualizado_en, a.creado_en), 'YYYY-MM-DD HH24:MI:SS') AS created_at,
          r.valor_likert AS valor
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND a.teacher_id = :tid
          AND r.question_id = :qid
          AND r.valor_likert IS NOT NULL
        ORDER BY COALESCE(r.created_at, a.actualizado_en, a.creado_en) DESC
    """), {"sid": str(survey_id), "tid": str(teacher_id), "qid": str(question_id)}).mappings().all()

    return QuestionTeacherDetailOut(
        question_id = qmeta["question_id"],
        codigo      = qmeta["codigo"],
        enunciado   = qmeta["enunciado"],
        section     = qmeta["section"],
        teacher_id  = trow["id"],
        teacher_nombre = trow["nombre"],
        n = n,
        avg = avg,
        dist = dist,
        attempts = [AttemptAnswerRow(**dict(r)) for r in attempts],
    )

# 4) DOCENTES (ranking + peor pregunta) + alias /teachers/summary
@router.get("/teachers", response_model=List[TeacherRowOut])
@router.get("/teachers/summary", response_model=List[TeacherRowOut])
def teachers_summary(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    q: Optional[str] = Query(None, description="Filtro ILIKE por nombre/identificador/programa"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)
    offset = (page - 1) * page_size

    rows = db.execute(text("""
        WITH base AS (
          SELECT a.teacher_id,
                 COUNT(DISTINCT a.id) AS n_respuestas,
                 AVG(r.valor_likert::numeric) AS promedio
          FROM public.attempts a
          JOIN public.responses r ON r.attempt_id = a.id
          WHERE a.survey_id = :sid
            AND a.estado = 'enviado'
            AND r.valor_likert IS NOT NULL
          GROUP BY a.teacher_id
        ),
        perq AS (
          SELECT a.teacher_id, r.question_id, AVG(r.valor_likert::numeric) AS avg_q
          FROM public.attempts a
          JOIN public.responses r ON r.attempt_id = a.id
          WHERE a.survey_id = :sid
            AND a.estado = 'enviado'
            AND r.valor_likert IS NOT NULL
          GROUP BY a.teacher_id, r.question_id
        ),
        peor AS (
          SELECT teacher_id, question_id, avg_q,
                 ROW_NUMBER() OVER (PARTITION BY teacher_id ORDER BY avg_q ASC) AS rn
          FROM perq
        )
        SELECT
          t.id             AS teacher_id,
          t.nombre         AS teacher_nombre,
          t.programa,
          COALESCE(b.n_respuestas, 0)   AS n_respuestas,
          b.promedio                     AS promedio,
          p.question_id                  AS peor_question_id,
          q.codigo                       AS peor_codigo,
          q.enunciado                    AS peor_enunciado,
          p.avg_q                        AS peor_promedio
        FROM public.survey_teacher_assignments sta
        JOIN public.teachers t ON t.id = sta.teacher_id
        LEFT JOIN base b ON b.teacher_id = t.id
        LEFT JOIN (SELECT teacher_id, question_id, avg_q FROM peor WHERE rn = 1) p ON p.teacher_id = t.id
        LEFT JOIN public.questions q ON q.id = p.question_id
        WHERE sta.survey_id = :sid
          AND (
            :q IS NULL OR
            t.nombre ILIKE '%' || :q || '%' OR
            t.identificador ILIKE '%' || :q || '%' OR
            COALESCE(t.programa,'') ILIKE '%' || :q || '%'
          )
        ORDER BY b.promedio DESC NULLS LAST, t.nombre
        LIMIT :limit OFFSET :offset
    """), {
        "sid": str(survey_id), "q": q, "limit": page_size, "offset": offset
    }).mappings().all()

    return [TeacherRowOut(**dict(r)) for r in rows]

# 5) MATRIZ DE CALOR 
@router.get("/teachers/matrix", response_model=TeacherMatrixOut)
def teachers_matrix(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    programa: Optional[str] = Query(None, description="Filtro por programa (ILIKE)"),
    min_n: int = Query(1, ge=1, description="Mínimo de respuestas por celda para reportar el promedio"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Matriz 'heatmap-ready': filas=docentes, columnas=códigos de pregunta (Q1..Qn),
    con el promedio por docente/pregunta. Si una celda tiene menos de `min_n` respuestas,
    se reporta como null.
    """
    _ensure_survey(db, survey_id)

    # 1) Columnas = códigos de pregunta (sólo tipo LIKERT; excluye 'texto'), orden por q.orden
    codes = db.execute(text("""
        SELECT q.codigo
        FROM public.questions q
        WHERE q.survey_id = :sid
          AND q.tipo <> 'texto'
        ORDER BY q.orden
    """), {"sid": str(survey_id)}).scalars().all()

    if not codes:
        return TeacherMatrixOut(columns=[], rows=[])

    # 2) Filas base = docentes asignados a la encuesta (+ n_respuestas)
    teachers = db.execute(text("""
        SELECT
          t.id   AS teacher_id,
          t.nombre AS teacher_nombre,
          t.programa,
          COUNT(DISTINCT a.id) FILTER (WHERE a.estado = 'enviado') AS n_respuestas
        FROM public.survey_teacher_assignments sta
        JOIN public.teachers t ON t.id = sta.teacher_id
        LEFT JOIN public.attempts a
               ON a.teacher_id = t.id AND a.survey_id = :sid
        WHERE sta.survey_id = :sid
          AND (:programa IS NULL OR t.programa ILIKE '%' || :programa || '%')
        GROUP BY t.id, t.nombre, t.programa
        ORDER BY t.nombre
    """), {"sid": str(survey_id), "programa": programa}).mappings().all()

    if not teachers:
        return TeacherMatrixOut(columns=codes, rows=[])

    # 3) Celdas: promedio y conteo por (docente, código de pregunta)
    cells = db.execute(text("""
        SELECT
          a.teacher_id,
          q.codigo,
          AVG(r.valor_likert::numeric) AS avg,
          COUNT(r.valor_likert)        AS n
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        JOIN public.questions q ON q.id = r.question_id
        WHERE a.survey_id = :sid
          AND a.estado   = 'enviado'
          AND r.valor_likert IS NOT NULL
          AND q.tipo <> 'texto'
          AND (
            :programa IS NULL OR EXISTS (
              SELECT 1 FROM public.teachers tt
              WHERE tt.id = a.teacher_id
                AND tt.programa ILIKE '%' || :programa || '%'
            )
          )
        GROUP BY a.teacher_id, q.codigo
    """), {"sid": str(survey_id), "programa": programa}).mappings().all()

    # 4) Pivot en Python: (teacher_id, codigo) -> (avg, n)
    cell_map = {}
    for r in cells:
        cell_map[(r["teacher_id"], r["codigo"])] = (
            float(r["avg"]) if r.get("avg") is not None else None,
            int(r["n"] or 0),
        )

    # 5) Construir filas en el mismo orden de "codes"
    rows_out: list[TeacherMatrixRow] = []
    for t in teachers:
        vals: list[Optional[float]] = []
        for code in codes:
            pair = cell_map.get((t["teacher_id"], code))
            if not pair:
                vals.append(None)
                continue
            avg, n = pair
            vals.append(avg if n >= min_n else None)

        rows_out.append(TeacherMatrixRow(
            teacher_id=t["teacher_id"],
            teacher_nombre=t["teacher_nombre"],
            programa=t["programa"],
            n_respuestas=int(t["n_respuestas"] or 0),
            values=vals,
        ))

    return TeacherMatrixOut(columns=codes, rows=rows_out)

@router.get("/teachers/filters", response_model=FiltersOut)
def teachers_filters(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    tz: str = Query("America/Bogota", description="Zona horaria para rango de fechas"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    # Pool de docentes: union de asignados y con intentos (por si no existe tabla de asignación en algún caso)
    pool_sql = text("""
        WITH pool AS (
          SELECT t.id, t.nombre, t.programa
          FROM public.teachers t
          JOIN public.survey_teacher_assignments sta ON sta.teacher_id = t.id
          WHERE sta.survey_id = :sid
          UNION
          SELECT t.id, t.nombre, t.programa
          FROM public.teachers t
          JOIN public.attempts a ON a.teacher_id = t.id
          WHERE a.survey_id = :sid
        )
        SELECT * FROM pool
    """)
    pool = db.execute(pool_sql, {"sid": str(survey_id)}).mappings().all()

    # Programas
    programas = sorted({(r.get("programa") or "").strip() for r in pool if r.get("programa")})  # únicos y no vacíos

    # Teachers
    teachers = [
        TeacherFilterItem(id=r["id"], nombre=r["nombre"], programa=r.get("programa"))
        for r in sorted(pool, key=lambda x: (x.get("nombre") or ""))
    ]

    # Sections
    sections_rows = db.execute(text("""
        SELECT s.id, s.titulo, COUNT(q.id) AS n_preguntas
        FROM public.survey_sections s
        LEFT JOIN public.questions q ON q.section_id = s.id
        WHERE s.survey_id = :sid
        GROUP BY s.id, s.titulo
        ORDER BY s.titulo
    """), {"sid": str(survey_id)}).mappings().all()

    sections = [
        SectionFilterItem(
            id=r["id"], titulo=r["titulo"], n_preguntas=int(r["n_preguntas"] or 0)
        ) for r in sections_rows
    ]

    # Questions (orden natural por q.orden)
    questions_rows = db.execute(text("""
        SELECT q.id, q.codigo, q.enunciado, s.titulo AS section
        FROM public.questions q
        JOIN public.survey_sections s ON s.id = q.section_id
        WHERE q.survey_id = :sid
        ORDER BY q.orden
    """), {"sid": str(survey_id)}).mappings().all()

    questions = [
        QuestionFilterItem(
            id=r["id"], codigo=r["codigo"], enunciado=r["enunciado"], section=r["section"]
        ) for r in questions_rows
    ]

    # Date range (enviados) – usando attempts.{creado_en,actualizado_en}; ajustado a tz
    dr = db.execute(text("""
        SELECT
          to_char(MIN((COALESCE(a.actualizado_en, a.creado_en)) AT TIME ZONE :tz), 'YYYY-MM-DD') AS min,
          to_char(MAX((COALESCE(a.actualizado_en, a.creado_en)) AT TIME ZONE :tz), 'YYYY-MM-DD') AS max
        FROM public.attempts a
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
    """), {"sid": str(survey_id), "tz": tz}).mappings().first() or {}

    date_range = DateRange(min=dr.get("min"), max=dr.get("max"))

    return FiltersOut(
        programas=programas,
        teachers=teachers,
        sections=sections,
        questions=questions,
        date_range=date_range,
    )


# 1) PREGUNTAS – estadísticas (mean/median/stddev, dist 1..5)
@router.get("/exports/questions-stats.csv")
def export_questions_stats_csv(
    survey_id: UUID = Query(...),
    min_n: int = Query(1, ge=0),
    include_ids: bool = Query(False),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    rows = db.execute(text("""
        SELECT
          q.id             AS question_id,
          q.codigo,
          q.enunciado,
          s.titulo         AS section,
          COUNT(r.valor_likert) AS n,
          AVG(r.valor_likert::numeric)                                        AS mean,
          PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY r.valor_likert)         AS median,
          STDDEV_POP(r.valor_likert::numeric)                                 AS stddev,
          MIN(r.valor_likert)                                                 AS min,
          MAX(r.valor_likert)                                                 AS max,
          SUM(CASE WHEN r.valor_likert = 1 THEN 1 ELSE 0 END)                 AS c1,
          SUM(CASE WHEN r.valor_likert = 2 THEN 1 ELSE 0 END)                 AS c2,
          SUM(CASE WHEN r.valor_likert = 3 THEN 1 ELSE 0 END)                 AS c3,
          SUM(CASE WHEN r.valor_likert = 4 THEN 1 ELSE 0 END)                 AS c4,
          SUM(CASE WHEN r.valor_likert = 5 THEN 1 ELSE 0 END)                 AS c5
        FROM public.responses r
        JOIN public.attempts a   ON a.id = r.attempt_id
        JOIN public.questions q  ON q.id = r.question_id
        JOIN public.survey_sections s ON s.id = q.section_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
        GROUP BY q.id, q.codigo, q.enunciado, s.titulo
        HAVING COUNT(r.valor_likert) >= :min_n
        ORDER BY q.orden
    """), {"sid": str(survey_id), "min_n": min_n}).mappings().all()

    def stream():
        output = io.StringIO()
        writer = csv.writer(output)
        headers = (["question_id"] if include_ids else []) + [
            "codigo","enunciado","section","n","mean","median","stddev","min","max","c1","c2","c3","c4","c5"
        ]
        writer.writerow(headers); yield output.getvalue(); output.seek(0); output.truncate(0)

        for r in rows:
            row = []
            if include_ids: row.append(str(r["question_id"]))
            row += [
                r["codigo"], r["enunciado"], r["section"],
                int(r["n"] or 0),
                float(r["mean"]) if r["mean"] is not None else None,
                float(r["median"]) if r["median"] is not None else None,
                float(r["stddev"]) if r["stddev"] is not None else None,
                int(r["min"]) if r["min"] is not None else None,
                int(r["max"]) if r["max"] is not None else None,
                int(r["c1"] or 0), int(r["c2"] or 0), int(r["c3"] or 0),
                int(r["c4"] or 0), int(r["c5"] or 0)
            ]
            writer.writerow(row); yield output.getvalue(); output.seek(0); output.truncate(0)

    filename = f"questions-stats_{survey_id}.csv"
    return StreamingResponse(stream(), media_type="text/csv",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# 2) DOCENTES – ranking + peor pregunta
@router.get("/exports/teachers-stats.csv")
def export_teachers_stats_csv(
    survey_id: UUID = Query(...),
    q: Optional[str] = Query(None, description="Filtro por nombre/identificador/programa"),
    include_ids: bool = Query(False),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    rows = db.execute(text("""
        WITH base AS (
          SELECT a.teacher_id,
                 COUNT(DISTINCT a.id)                    AS n_respuestas,
                 AVG(r.valor_likert::numeric)            AS promedio
          FROM public.attempts a
          JOIN public.responses r ON r.attempt_id = a.id
          WHERE a.survey_id = :sid
            AND a.estado = 'enviado'
            AND r.valor_likert IS NOT NULL
          GROUP BY a.teacher_id
        ),
        perq AS (
          SELECT a.teacher_id, r.question_id, AVG(r.valor_likert::numeric) AS avg_q
          FROM public.attempts a
          JOIN public.responses r ON r.attempt_id = a.id
          WHERE a.survey_id = :sid
            AND a.estado = 'enviado'
            AND r.valor_likert IS NOT NULL
          GROUP BY a.teacher_id, r.question_id
        ),
        peor AS (
          SELECT teacher_id, question_id, avg_q,
                 ROW_NUMBER() OVER (PARTITION BY teacher_id ORDER BY avg_q ASC) AS rn
          FROM perq
        )
        SELECT
          t.id             AS teacher_id,
          t.identificador  AS docente_identificador,
          t.nombre         AS teacher_nombre,
          t.programa,
          COALESCE(b.n_respuestas, 0)   AS n_respuestas,
          b.promedio                     AS promedio,
          p.question_id                  AS peor_question_id,
          q.codigo                       AS peor_codigo,
          q.enunciado                    AS peor_enunciado,
          p.avg_q                        AS peor_promedio
        FROM public.survey_teacher_assignments sta
        JOIN public.teachers t ON t.id = sta.teacher_id
        LEFT JOIN base b ON b.teacher_id = t.id
        LEFT JOIN (SELECT teacher_id, question_id, avg_q FROM peor WHERE rn = 1) p ON p.teacher_id = t.id
        LEFT JOIN public.questions q ON q.id = p.question_id
        WHERE sta.survey_id = :sid
          AND (
            :q IS NULL OR
            t.nombre ILIKE '%' || :q || '%' OR
            t.identificador ILIKE '%' || :q || '%' OR
            COALESCE(t.programa,'') ILIKE '%' || :q || '%'
          )
        ORDER BY b.promedio DESC NULLS LAST, t.nombre
    """), {"sid": str(survey_id), "q": q}).mappings().all()

    def stream():
        output = io.StringIO()
        writer = csv.writer(output)
        headers = (["teacher_id"] if include_ids else []) + [
            "docente_identificador","teacher_nombre","programa",
            "n_respuestas","promedio","peor_codigo","peor_enunciado","peor_promedio"
        ]
        writer.writerow(headers); yield output.getvalue(); output.seek(0); output.truncate(0)

        for r in rows:
            row = []
            if include_ids: row.append(str(r["teacher_id"]))
            row += [
                r["docente_identificador"], r["teacher_nombre"], r.get("programa"),
                int(r["n_respuestas"] or 0),
                float(r["promedio"]) if r.get("promedio") is not None else None,
                r.get("peor_codigo"), r.get("peor_enunciado"),
                float(r["peor_promedio"]) if r.get("peor_promedio") is not None else None
            ]
            writer.writerow(row); yield output.getvalue(); output.seek(0); output.truncate(0)

    filename = f"teachers-stats_{survey_id}.csv"
    return StreamingResponse(stream(), media_type="text/csv",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# 3) MATRIZ – heatmap ready (filtrable por programa)
@router.get("/exports/matrix.csv")
def export_matrix_csv(
    survey_id: UUID = Query(...),
    programa: Optional[str] = Query(None),
    min_n: int = Query(1, ge=0),
    include_ids: bool = Query(False),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    # 1) Columnas Q* (no texto) en orden
    qcodes = db.execute(text("""
        SELECT codigo
        FROM public.questions
        WHERE survey_id = :sid AND tipo <> 'texto'
        ORDER BY orden
    """), {"sid": str(survey_id)}).scalars().all()

    # 2) SELECT dinámico pivotado desde la CTE perq (pq)
    cols_sql = []
    for code in qcodes:
        safe = "".join(ch for ch in code if ch.isalnum() or ch == "_")
        cols_sql.append(
            f'AVG(CASE WHEN pq.codigo = \'{safe}\' THEN pq.avg_q END) AS "{safe}"'
        )
    cols_sql_joined = ",\n          ".join(cols_sql) if cols_sql else "NULL"

    # 3) Query
    sql = f"""
        WITH base AS (
          SELECT a.teacher_id,
                 COUNT(DISTINCT a.id) AS n_respuestas
          FROM public.attempts a
          WHERE a.survey_id = :sid AND a.estado = 'enviado'
          GROUP BY a.teacher_id
        ),
        perq AS (
          SELECT a.teacher_id, q.codigo, AVG(r.valor_likert::numeric) AS avg_q
          FROM public.responses r
          JOIN public.attempts a ON a.id = r.attempt_id
          JOIN public.questions q ON q.id = r.question_id
          WHERE a.survey_id = :sid
            AND a.estado = 'enviado'
            AND r.valor_likert IS NOT NULL
          GROUP BY a.teacher_id, q.codigo
        )
        SELECT
          t.id AS teacher_id,
          t.nombre AS teacher_nombre,
          t.programa,
          COALESCE(b.n_respuestas, 0) AS n_respuestas,
          {cols_sql_joined}
        FROM public.survey_teacher_assignments sta
        JOIN public.teachers t ON t.id = sta.teacher_id
        LEFT JOIN base b ON b.teacher_id = t.id
        LEFT JOIN perq pq ON pq.teacher_id = t.id
        WHERE sta.survey_id = :sid
          {"AND t.programa = :programa" if programa else ""}
        GROUP BY t.id, t.nombre, t.programa, b.n_respuestas
        HAVING COALESCE(b.n_respuestas, 0) >= :min_n
        ORDER BY t.nombre
    """

    params = {"sid": str(survey_id), "min_n": min_n}
    if programa:
        params["programa"] = programa

    rows = db.execute(text(sql), params).mappings().all()

    def stream():
        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        headers = (["teacher_id"] if include_ids else []) + ["teacher_nombre","programa","n_respuestas"] + qcodes
        writer.writerow(headers); yield output.getvalue(); output.seek(0); output.truncate(0)

        for r in rows:
            row = []
            if include_ids: row.append(str(r["teacher_id"]))
            row += [r["teacher_nombre"], r.get("programa"), int(r["n_respuestas"] or 0)]
            for code in qcodes:
                val = r.get(code)
                row.append(float(val) if val is not None else None)
            writer.writerow(row); yield output.getvalue(); output.seek(0); output.truncate(0)

    filename = f"matrix_{survey_id}.csv"
    return StreamingResponse(stream(), media_type="text/csv",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.get("/exports/survey/{survey_id}/teachers.csv")
def export_survey_teachers_csv(
    survey_id: UUID,
    min_n: int = Query(1, ge=0),
    include_ids: bool = Query(False),
    programa: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    sql = """
    WITH base AS (
      SELECT a.teacher_id,
             COUNT(DISTINCT a.id) AS n_respuestas,
             AVG(r.valor_likert::numeric) AS promedio_global
      FROM public.attempts a
      JOIN public.responses r ON r.attempt_id = a.id
      JOIN public.questions q ON q.id = r.question_id
      WHERE a.survey_id = :sid
        AND a.estado = 'enviado'
        AND r.valor_likert IS NOT NULL
        AND q.tipo <> 'texto'
      GROUP BY a.teacher_id
    ),
    perq AS (
      SELECT a.teacher_id, q.id AS question_id, q.codigo, q.enunciado,
             AVG(r.valor_likert::numeric) AS avg_q
      FROM public.attempts a
      JOIN public.responses r ON r.attempt_id = a.id
      JOIN public.questions q ON q.id = r.question_id
      WHERE a.survey_id = :sid
        AND a.estado = 'enviado'
        AND r.valor_likert IS NOT NULL
        AND q.tipo <> 'texto'
      GROUP BY a.teacher_id, q.id, q.codigo, q.enunciado
    ),
    worst AS (
      SELECT DISTINCT ON (teacher_id)
             teacher_id, question_id, codigo, enunciado, avg_q
      FROM perq
      ORDER BY teacher_id, avg_q ASC, codigo
    )
    SELECT
      ROW_NUMBER() OVER (ORDER BY b.promedio_global DESC NULLS LAST, t.nombre) AS ranking,
      t.id AS teacher_id,
      t.identificador AS docente_identificador,
      t.nombre AS docente_nombre,
      t.programa AS docente_programa,
      COALESCE(b.n_respuestas, 0) AS n_respuestas,
      b.promedio_global,
      w.question_id AS peor_question_id,
      w.codigo AS peor_codigo,
      w.enunciado AS peor_enunciado,
      w.avg_q AS peor_promedio
    FROM public.survey_teacher_assignments sta
    JOIN public.teachers t ON t.id = sta.teacher_id
    LEFT JOIN base b   ON b.teacher_id = t.id
    LEFT JOIN worst w  ON w.teacher_id = t.id
    WHERE sta.survey_id = :sid
      {prog_filter}
      AND COALESCE(b.n_respuestas, 0) >= :min_n
    ORDER BY ranking
    """.format(prog_filter="AND t.programa = :programa" if programa else "")

    params = {"sid": str(survey_id), "min_n": min_n}
    if programa:
        params["programa"] = programa

    rows = db.execute(text(sql), params).mappings().all()

    import csv, io
    def stream():
        out = io.StringIO()
        w = csv.writer(out)
        headers = (["teacher_id"] if include_ids else []) + [
            "ranking","docente_identificador","docente_nombre","docente_programa",
            "n_respuestas","promedio_global",
            "peor_codigo","peor_enunciado","peor_promedio"
        ]
        w.writeheader if hasattr(w, "writeheader") else None
        w.writerow(headers); yield out.getvalue(); out.seek(0); out.truncate(0)

        for r in rows:
            row = []
            if include_ids: row.append(str(r["teacher_id"]))
            row += [
                int(r["ranking"]),
                r["docente_identificador"],
                r["docente_nombre"],
                r["docente_programa"],
                int(r["n_respuestas"] or 0),
                (float(r["promedio_global"]) if r["promedio_global"] is not None else None),
                r.get("peor_codigo"),
                r.get("peor_enunciado"),
                (float(r["peor_promedio"]) if r.get("peor_promedio") is not None else None),
            ]
            w.writerow(row); yield out.getvalue(); out.seek(0); out.truncate(0)

    filename = f"survey_{survey_id}_teachers.csv"
    return StreamingResponse(stream(), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.get("/exports/survey/{survey_id}/comments.csv")
def export_survey_comments_csv(
    survey_id: UUID,
    tz: str = Query("America/Bogota"),
    include_ids: bool = Query(False),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    sql = """
    SELECT
      a.id AS attempt_id,
      t.identificador AS docente_identificador,
      t.nombre AS docente_nombre,
      t.programa AS docente_programa,
      u.email AS usuario_email,
      COALESCE(u.nombre, u.email) AS usuario_nombre,
      to_char((COALESCE(a.actualizado_en, a.creado_en)) AT TIME ZONE :tz,
              'YYYY-MM-DD HH24:MI:SS') AS enviado_local,
      r.texto->>'positivos'   AS positivos,
      r.texto->>'mejorar'     AS mejorar,
      r.texto->>'comentarios' AS comentarios
    FROM public.attempts a
      JOIN public.responses r ON r.attempt_id = a.id
      JOIN public.questions q ON q.id = r.question_id
      LEFT JOIN public.users u ON u.id = a.user_id
      JOIN public.teachers t ON t.id = a.teacher_id
    WHERE a.survey_id = :sid
      AND a.estado = 'enviado'
      AND q.tipo = 'texto'
    ORDER BY enviado_local DESC, docente_nombre
    """

    rows = db.execute(text(sql), {"sid": str(survey_id), "tz": tz}).mappings().all()

    import csv, io
    def stream():
        out = io.StringIO()
        w = csv.writer(out)
        headers = (["attempt_id"] if include_ids else []) + [
            "docente_identificador","docente_nombre","docente_programa",
            "usuario_email","usuario_nombre","enviado_local",
            "positivos","mejorar","comentarios"
        ]
        w.writerow(headers); yield out.getvalue(); out.seek(0); out.truncate(0)

        for r in rows:
            row = []
            if include_ids: row.append(str(r["attempt_id"]))
            row += [
                r["docente_identificador"], r["docente_nombre"], r["docente_programa"],
                r["usuario_email"], r["usuario_nombre"], r["enviado_local"],
                r.get("positivos"), r.get("mejorar"), r.get("comentarios")
            ]
            w.writerow(row); yield out.getvalue(); out.seek(0); out.truncate(0)

    filename = f"survey_{survey_id}_comments.csv"
    return StreamingResponse(stream(), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@router.get("/exports/survey/{survey_id}.xlsx")
def export_survey_xlsx(
    survey_id: UUID,
    tz: str = Query("America/Bogota"),
    min_n: int = Query(1, ge=0),
    programa: Optional[str] = Query(None, description="Opcional: filtra docentes por programa en la hoja Docentes"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    # ---------- Queries ----------

    # Global y por secciones (para 'Resumen' y 'Secciones')
    q_resumen = text("""
      WITH global AS (
        SELECT
          COUNT(DISTINCT a.id) AS n_intentos,
          AVG(r.valor_likert::numeric) AS promedio_global
        FROM public.attempts a
        JOIN public.responses r ON r.attempt_id = a.id
        JOIN public.questions q ON q.id = r.question_id
        WHERE a.survey_id = :sid AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL AND q.tipo <> 'texto'
      ),
      secciones AS (
        SELECT s.id, s.titulo,
               COUNT(r.*) AS n_respuestas,
               AVG(r.valor_likert::numeric) AS promedio
        FROM public.survey_sections s
        JOIN public.questions q ON q.section_id = s.id AND q.survey_id = :sid AND q.tipo <> 'texto'
        JOIN public.responses r ON r.question_id = q.id
        JOIN public.attempts a ON a.id = r.attempt_id AND a.estado = 'enviado' AND a.survey_id = :sid
        GROUP BY s.id, s.titulo
        ORDER BY s.titulo
      )
      SELECT * FROM global;
    """)
    resumen = db.execute(q_resumen, {"sid": str(survey_id)}).mappings().first()

    q_secciones = text("""
      SELECT s.titulo, COUNT(r.*) AS n_respuestas, AVG(r.valor_likert::numeric) AS promedio
      FROM public.survey_sections s
      JOIN public.questions q ON q.section_id = s.id AND q.survey_id = :sid AND q.tipo <> 'texto'
      JOIN public.responses r ON r.question_id = q.id
      JOIN public.attempts a ON a.id = r.attempt_id AND a.estado = 'enviado' AND a.survey_id = :sid
      GROUP BY s.titulo
      ORDER BY s.titulo
    """)
    secciones = db.execute(q_secciones, {"sid": str(survey_id)}).mappings().all()

    # Preguntas: n, mean, median, stddev, c1..c5
    q_preg = text("""
      SELECT
        q.codigo, q.enunciado,
        COUNT(r.*) AS n,
        AVG(r.valor_likert::numeric) AS mean,
        PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY r.valor_likert) AS median,
        STDDEV_SAMP(r.valor_likert::numeric) AS stddev,
        SUM(CASE WHEN r.valor_likert=1 THEN 1 ELSE 0 END) AS c1,
        SUM(CASE WHEN r.valor_likert=2 THEN 1 ELSE 0 END) AS c2,
        SUM(CASE WHEN r.valor_likert=3 THEN 1 ELSE 0 END) AS c3,
        SUM(CASE WHEN r.valor_likert=4 THEN 1 ELSE 0 END) AS c4,
        SUM(CASE WHEN r.valor_likert=5 THEN 1 ELSE 0 END) AS c5
      FROM public.questions q
      JOIN public.responses r ON r.question_id = q.id
      JOIN public.attempts a ON a.id = r.attempt_id
      WHERE q.survey_id = :sid
        AND q.tipo <> 'texto'
        AND a.estado = 'enviado'
        AND r.valor_likert IS NOT NULL
      GROUP BY q.codigo, q.enunciado
      ORDER BY q.codigo
    """)
    preguntas = db.execute(q_preg, {"sid": str(survey_id)}).mappings().all()

    # Docentes (ranking + peor pregunta)
    q_doc = text("""
      WITH base AS (
        SELECT a.teacher_id,
               COUNT(DISTINCT a.id) AS n_respuestas,
               AVG(r.valor_likert::numeric) AS promedio_global
        FROM public.attempts a
        JOIN public.responses r ON r.attempt_id = a.id
        JOIN public.questions q ON q.id = r.question_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
          AND q.tipo <> 'texto'
        GROUP BY a.teacher_id
      ),
      perq AS (
        SELECT a.teacher_id, q.id AS question_id, q.codigo, q.enunciado,
               AVG(r.valor_likert::numeric) AS avg_q
        FROM public.attempts a
        JOIN public.responses r ON r.attempt_id = a.id
        JOIN public.questions q ON q.id = r.question_id
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
          AND q.tipo <> 'texto'
        GROUP BY a.teacher_id, q.id, q.codigo, q.enunciado
      ),
      worst AS (
        SELECT DISTINCT ON (teacher_id)
               teacher_id, question_id, codigo, enunciado, avg_q
        FROM perq
        ORDER BY teacher_id, avg_q ASC, codigo
      )
      SELECT
        ROW_NUMBER() OVER (ORDER BY b.promedio_global DESC NULLS LAST, t.nombre) AS ranking,
        t.identificador AS docente_identificador,
        t.nombre AS docente_nombre,
        t.programa AS docente_programa,
        COALESCE(b.n_respuestas, 0) AS n_respuestas,
        b.promedio_global,
        w.codigo AS peor_codigo,
        w.enunciado AS peor_enunciado,
        w.avg_q AS peor_promedio
      FROM public.survey_teacher_assignments sta
      JOIN public.teachers t ON t.id = sta.teacher_id
      LEFT JOIN base b   ON b.teacher_id = t.id
      LEFT JOIN worst w  ON w.teacher_id = t.id
      WHERE sta.survey_id = :sid
        {prog_filter}
        AND COALESCE(b.n_respuestas, 0) >= :min_n
      ORDER BY ranking
    """.replace("{prog_filter}", "AND t.programa = :programa" if programa else ""))
    p = {"sid": str(survey_id), "min_n": min_n}
    if programa: p["programa"] = programa
    docentes = db.execute(q_doc, p).mappings().all()

    # Comentarios
    q_com = text("""
      SELECT
        t.identificador AS docente_identificador,
        t.nombre AS docente_nombre,
        t.programa AS docente_programa,
        u.email AS usuario_email,
        COALESCE(u.nombre, u.email) AS usuario_nombre,
        to_char((COALESCE(a.actualizado_en, a.creado_en)) AT TIME ZONE :tz,
                'YYYY-MM-DD HH24:MI:SS') AS enviado_local,
        r.texto->>'positivos'   AS positivos,
        r.texto->>'mejorar'     AS mejorar,
        r.texto->>'comentarios' AS comentarios
      FROM public.attempts a
      JOIN public.responses r ON r.attempt_id = a.id
      JOIN public.questions q ON q.id = r.question_id
      LEFT JOIN public.users u ON u.id = a.user_id
      JOIN public.teachers t ON t.id = a.teacher_id
      WHERE a.survey_id = :sid AND a.estado = 'enviado' AND q.tipo = 'texto'
      ORDER BY enviado_local DESC, docente_nombre
    """)
    comentarios = db.execute(q_com, {"sid": str(survey_id), "tz": tz}).mappings().all()

    # Progreso diario
    q_prog = text("""
      SELECT
        to_char(date_trunc('day', (COALESCE(a.actualizado_en, a.creado_en)) AT TIME ZONE 'UTC'), 'YYYY-MM-DD') AS day,
        COUNT(*) AS sent
      FROM public.attempts a
      WHERE a.survey_id = :sid AND a.estado = 'enviado'
      GROUP BY 1
      ORDER BY 1
    """)
    progreso = db.execute(q_prog, {"sid": str(survey_id)}).mappings().all()

    # ---------- Excel ----------
    wb = Workbook()
    ws_res = wb.active; ws_res.title = "Resumen"

    # Resumen (fila simple)
    ws_res.append(["n_intentos", "promedio_global"])
    ws_res.append([
        int(resumen["n_intentos"] or 0) if resumen else 0,
        float(resumen["promedio_global"]) if resumen and resumen["promedio_global"] is not None else None
    ])

    # Secciones
    ws_sec = wb.create_sheet("Secciones")
    ws_sec.append(["seccion","n_respuestas","promedio"])
    for s in secciones:
        ws_sec.append([s["titulo"], int(s["n_respuestas"] or 0),
                       float(s["promedio"]) if s["promedio"] is not None else None])

    # Preguntas
    ws_q = wb.create_sheet("Preguntas")
    ws_q.append(["codigo","enunciado","n","mean","median","stddev","c1","c2","c3","c4","c5"])
    for r in preguntas:
        ws_q.append([
            r["codigo"], r["enunciado"], int(r["n"] or 0),
            float(r["mean"]) if r["mean"] is not None else None,
            float(r["median"]) if r["median"] is not None else None,
            float(r["stddev"]) if r["stddev"] is not None else None,
            int(r["c1"] or 0), int(r["c2"] or 0), int(r["c3"] or 0), int(r["c4"] or 0), int(r["c5"] or 0)
        ])

    # Docentes
    ws_t = wb.create_sheet("Docentes")
    ws_t.append(["ranking","docente_identificador","docente_nombre","docente_programa",
                 "n_respuestas","promedio_global","peor_codigo","peor_enunciado","peor_promedio"])
    for r in docentes:
        ws_t.append([
            int(r["ranking"]),
            r["docente_identificador"], r["docente_nombre"], r["docente_programa"],
            int(r["n_respuestas"] or 0),
            float(r["promedio_global"]) if r["promedio_global"] is not None else None,
            r.get("peor_codigo"), r.get("peor_enunciado"),
            float(r["peor_promedio"]) if r.get("peor_promedio") is not None else None
        ])

    # Comentarios
    ws_c = wb.create_sheet("Comentarios")
    ws_c.append(["docente_identificador","docente_nombre","docente_programa",
                 "usuario_email","usuario_nombre","enviado_local",
                 "positivos","mejorar","comentarios"])
    for r in comentarios:
        ws_c.append([
            r["docente_identificador"], r["docente_nombre"], r["docente_programa"],
            r["usuario_email"], r["usuario_nombre"], r["enviado_local"],
            r.get("positivos"), r.get("mejorar"), r.get("comentarios")
        ])

    # Progreso
    ws_p = wb.create_sheet("Progreso")
    ws_p.append(["day","sent"])
    for r in progreso:
        ws_p.append([r["day"], int(r["sent"] or 0)])

    # Guardar en memoria y responder
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f"survey_{survey_id}.xlsx"
    return StreamingResponse(iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})



# 5) DETALLE DOCENTE
@router.get("/teachers/{teacher_id}", response_model=TeacherDetailOut)
def teacher_detail(
    teacher_id: UUID = Path(...),
    survey_id: UUID = Query(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    head = db.execute(text("""
        SELECT t.id AS teacher_id, t.nombre AS teacher_nombre, t.programa,
               COUNT(DISTINCT a.id) AS n_respuestas,
               AVG(r.valor_likert::numeric) AS promedio
        FROM public.teachers t
        JOIN public.attempts a ON a.teacher_id = t.id
        LEFT JOIN public.responses r ON r.attempt_id = a.id AND r.valor_likert IS NOT NULL
        WHERE a.survey_id = :sid AND a.estado = 'enviado' AND t.id = :tid
        GROUP BY t.id, t.nombre, t.programa
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().first()

    if not head:
        head = db.execute(text("""
            SELECT t.id AS teacher_id, t.nombre AS teacher_nombre, t.programa
            FROM public.teachers t
            JOIN public.survey_teacher_assignments sta ON sta.teacher_id = t.id
            WHERE sta.survey_id = :sid AND t.id = :tid
        """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().first()
        if not head:
            raise HTTPException(404, "Docente no encontrado en la encuesta")
        n_respuestas = 0
        promedio = None
    else:
        n_respuestas = int(head.get("n_respuestas") or 0)
        promedio = float(head.get("promedio")) if head.get("promedio") is not None else None

    qrows = db.execute(text("""
        SELECT q.id AS question_id, q.codigo, q.enunciado, s.titulo AS section,
               COUNT(r.valor_likert) AS n,
               AVG(r.valor_likert::numeric) AS mean,
               SUM(CASE WHEN r.valor_likert = 1 THEN 1 ELSE 0 END) AS c1,
               SUM(CASE WHEN r.valor_likert = 2 THEN 1 ELSE 0 END) AS c2,
               SUM(CASE WHEN r.valor_likert = 3 THEN 1 ELSE 0 END) AS c3,
               SUM(CASE WHEN r.valor_likert = 4 THEN 1 ELSE 0 END) AS c4,
               SUM(CASE WHEN r.valor_likert = 5 THEN 1 ELSE 0 END) AS c5
        FROM public.responses r
        JOIN public.attempts a   ON a.id = r.attempt_id
        JOIN public.questions q  ON q.id = r.question_id
        JOIN public.survey_sections s ON s.id = q.section_id
        WHERE a.survey_id = :sid
          AND a.teacher_id = :tid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
        GROUP BY q.id, q.codigo, q.enunciado, s.titulo
        ORDER BY q.orden
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().all()

    preguntas = [TeacherQBreakdown(**dict(r)) for r in qrows]

    comments = db.execute(text("""
        SELECT a.id AS attempt_id,
               to_char(COALESCE(r.created_at, a.creado_en), 'YYYY-MM-DD HH24:MI:SS') AS creado_en,
               r.texto->>'positivos'   AS positivos,
               r.texto->>'mejorar'     AS mejorar,
               r.texto->>'comentarios' AS comentarios
        FROM public.attempts a
        JOIN public.responses r ON r.attempt_id = a.id
        JOIN public.questions q ON q.id = r.question_id
        WHERE a.survey_id = :sid
          AND a.teacher_id = :tid
          AND a.estado = 'enviado'
          AND (q.codigo = 'Q16' OR q.tipo = 'texto')
          AND r.texto IS NOT NULL
        ORDER BY COALESCE(r.created_at, a.creado_en) DESC
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().all()

    return TeacherDetailOut(
        teacher_id=head["teacher_id"],
        teacher_nombre=head["teacher_nombre"],
        programa=head.get("programa"),
        n_respuestas=n_respuestas,
        promedio=promedio,
        preguntas=preguntas,
        comentarios=[CommentRow(**dict(r)) for r in comments],
    )

# 5b) PROMEDIOS POR SECCIÓN PARA UN DOCENTE
@router.get("/teachers/{teacher_id}/sections", response_model=TeacherSectionsOut)
def teacher_sections(
    teacher_id: UUID = Path(...),
    survey_id: UUID = Query(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Retorna los promedios por sección para un docente específico.
    Solo considera respuestas 'enviadas' de tipo likert.
    """
    _ensure_survey(db, survey_id)
    
    # Verificar que el docente existe y está asignado
    teacher = db.execute(text("""
        SELECT t.id, t.nombre
        FROM public.teachers t
        JOIN public.survey_teacher_assignments sta ON sta.teacher_id = t.id
        WHERE sta.survey_id = :sid AND t.id = :tid
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().first()
    
    if not teacher:
        raise HTTPException(404, "Docente no encontrado en esta encuesta")
    
    # Obtener promedios por sección
    rows = db.execute(text("""
        SELECT 
            s.id AS section_id,
            s.titulo,
            COUNT(r.valor_likert) AS n_respuestas,
            AVG(r.valor_likert::numeric) AS promedio
        FROM public.survey_sections s
        JOIN public.questions q ON q.section_id = s.id
        JOIN public.responses r ON r.question_id = q.id
        JOIN public.attempts a ON a.id = r.attempt_id
        WHERE a.survey_id = :sid
          AND a.teacher_id = :tid
          AND a.estado = 'enviado'
          AND r.valor_likert IS NOT NULL
          AND q.tipo = 'likert'
        GROUP BY s.id, s.titulo
        ORDER BY s.titulo
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().all()
    
    sections = [
        TeacherSectionScore(
            section_id=r["section_id"],
            titulo=r["titulo"],
            n_respuestas=int(r["n_respuestas"] or 0),
            promedio=float(r["promedio"]) if r.get("promedio") is not None else None
        )
        for r in rows
    ]
    
    return TeacherSectionsOut(
        teacher_id=teacher["id"],
        teacher_nombre=teacher["nombre"],
        sections=sections
    )

# 5c) MAPA DE CALOR DE ESTUDIANTES PARA UN DOCENTE
@router.get("/teachers/{teacher_id}/students-heatmap", response_model=StudentHeatmapOut)
def teacher_students_heatmap(
    teacher_id: UUID = Path(...),
    survey_id: UUID = Query(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Mapa de calor de respuestas por estudiante (cada intento/attempt) para un docente.
    Cada fila representa un intento (estudiante) con sus respuestas a cada pregunta.
    Solo considera respuestas 'enviadas' de tipo likert.
    """
    _ensure_survey(db, survey_id)
    
    # Verificar que el docente existe
    teacher = db.execute(text("""
        SELECT t.id, t.nombre
        FROM public.teachers t
        JOIN public.survey_teacher_assignments sta ON sta.teacher_id = t.id
        WHERE sta.survey_id = :sid AND t.id = :tid
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().first()
    
    if not teacher:
        raise HTTPException(404, "Docente no encontrado en esta encuesta")
    
    # 1) Obtener columnas (códigos de preguntas ordenadas)
    codes = db.execute(text("""
        SELECT q.codigo
        FROM public.questions q
        WHERE q.survey_id = :sid
          AND q.tipo <> 'texto'
        ORDER BY q.orden
    """), {"sid": str(survey_id)}).scalars().all()
    
    if not codes:
        return StudentHeatmapOut(
            teacher_id=teacher["id"],
            teacher_nombre=teacher["nombre"],
            columns=[],
            rows=[]
        )
    
    # 2) Obtener intentos (estudiantes) para este docente
    attempts = db.execute(text("""
        SELECT 
            a.id AS attempt_id,
            u.email AS user_email,
            to_char(COALESCE(a.actualizado_en, a.creado_en), 'YYYY-MM-DD HH24:MI:SS') AS created_at,
            COUNT(r.valor_likert) AS n_respuestas,
            AVG(r.valor_likert::numeric) AS promedio
        FROM public.attempts a
        LEFT JOIN public.users u ON u.id = a.user_id
        LEFT JOIN public.responses r ON r.attempt_id = a.id AND r.valor_likert IS NOT NULL
        LEFT JOIN public.questions q ON q.id = r.question_id AND q.tipo <> 'texto'
        WHERE a.survey_id = :sid
          AND a.teacher_id = :tid
          AND a.estado = 'enviado'
        GROUP BY a.id, u.email, a.actualizado_en, a.creado_en
        ORDER BY COALESCE(a.actualizado_en, a.creado_en) DESC
    """), {"sid": str(survey_id), "tid": str(teacher_id)}).mappings().all()
    
    if not attempts:
        return StudentHeatmapOut(
            teacher_id=teacher["id"],
            teacher_nombre=teacher["nombre"],
            columns=codes,
            rows=[]
        )
    
    # 3) Obtener todas las respuestas para estos intentos
    attempt_ids_list = [a["attempt_id"] for a in attempts]
    
    if not attempt_ids_list:
        cells = []
    else:
        # Crear placeholders dinámicos para los UUIDs
        placeholders = ",".join([f":attempt_id_{i}" for i in range(len(attempt_ids_list))])
        params = {f"attempt_id_{i}": attempt_ids_list[i] for i in range(len(attempt_ids_list))}
        
        cells = db.execute(text(f"""
            SELECT 
                r.attempt_id,
                q.codigo,
                r.valor_likert::numeric AS valor
            FROM public.responses r
            JOIN public.questions q ON q.id = r.question_id
            WHERE r.attempt_id IN ({placeholders})
              AND r.valor_likert IS NOT NULL
              AND q.tipo <> 'texto'
        """), params).mappings().all()
    
    # 4) Crear mapa (attempt_id, codigo) -> valor
    cell_map = {}
    for r in cells:
        cell_map[(r["attempt_id"], r["codigo"])] = float(r["valor"]) if r.get("valor") is not None else None
    
    # 5) Construir filas
    rows_out = []
    for att in attempts:
        vals = []
        for code in codes:
            val = cell_map.get((att["attempt_id"], code))
            vals.append(val)
        
        rows_out.append(StudentHeatmapRow(
            attempt_id=att["attempt_id"],
            user_email=att.get("user_email"),
            created_at=att.get("created_at"),
            n_respuestas=int(att["n_respuestas"] or 0),
            promedio=float(att["promedio"]) if att.get("promedio") is not None else None,
            values=vals
        ))
    
    return StudentHeatmapOut(
        teacher_id=teacher["id"],
        teacher_nombre=teacher["nombre"],
        columns=codes,
        rows=rows_out
    )

# 6) COMMENTS (Q16) – búsqueda/paginación
@router.get("/comments", response_model=CommentListOut)
def list_comments(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    teacher_id: Optional[UUID] = Query(None, description="Filtra por docente"),
    q: Optional[str] = Query(None, description="Texto libre en Q16 (positivos/mejorar/comentarios)"),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    _ensure_survey(db, survey_id)

    params = {"sid": str(survey_id), "tid": str(teacher_id) if teacher_id else None,
              "qq": q, "limit": limit, "offset": offset}

    total_row = db.execute(text("""
        SELECT COUNT(*) AS total
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        JOIN public.questions qn ON qn.id = r.question_id
        WHERE a.survey_id = :sid
          AND a.estado   = 'enviado'
          AND r.texto IS NOT NULL
          AND (qn.codigo = 'Q16' OR qn.tipo = 'texto')
          AND (:tid IS NULL OR a.teacher_id = :tid)
          AND (
            :qq IS NULL OR
            (r.texto->>'positivos'   ILIKE '%' || :qq || '%'
             OR r.texto->>'mejorar'  ILIKE '%' || :qq || '%'
             OR r.texto->>'comentarios' ILIKE '%' || :qq || '%')
          )
    """), params).mappings().first()
    total = int(total_row["total"] or 0)

    rows = db.execute(text("""
        SELECT a.id AS attempt_id,
               a.teacher_id,
               t.nombre AS teacher_nombre,
               to_char(COALESCE(r.created_at, a.creado_en), 'YYYY-MM-DD HH24:MI:SS') AS created_at,
               r.texto->>'positivos'   AS positivos,
               r.texto->>'mejorar'     AS mejorar,
               r.texto->>'comentarios' AS comentarios
        FROM public.responses r
        JOIN public.attempts a ON a.id = r.attempt_id
        JOIN public.teachers t ON t.id = a.teacher_id
        JOIN public.questions qn ON qn.id = r.question_id
        WHERE a.survey_id = :sid
          AND a.estado   = 'enviado'
          AND r.texto IS NOT NULL
          AND (qn.codigo = 'Q16' OR qn.tipo = 'texto')
          AND (:tid IS NULL OR a.teacher_id = :tid)
          AND (
            :qq IS NULL OR
            (r.texto->>'positivos'   ILIKE '%' || :qq || '%'
             OR r.texto->>'mejorar'  ILIKE '%' || :qq || '%'
             OR r.texto->>'comentarios' ILIKE '%' || :qq || '%')
          )
        ORDER BY COALESCE(r.created_at, a.creado_en) DESC
        LIMIT :limit OFFSET :offset
    """), params).mappings().all()

    items = [CommentListItem(**dict(r)) for r in rows]
    return CommentListOut(total=total, items=items)
# =============================
# 7) PROGRESO DIARIO (serie temporal)
# =============================
@router.get("/progress/daily", response_model=ProgressDailyOut)
def progress_daily(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    date_from: Optional[str] = Query(None, alias="from", description="YYYY-MM-DD (opcional)"),
    date_to: Optional[str] = Query(None, alias="to", description="YYYY-MM-DD (opcional)"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Conteo de intentos 'enviados' por día.
    Usa COALESCE(a.actualizado_en, a.creado_en) como marca temporal.
    """
    _ensure_survey(db, survey_id)

    rows = db.execute(text("""
        SELECT
          to_char(
            date_trunc('day',
              COALESCE(a.actualizado_en, a.creado_en) AT TIME ZONE 'UTC'
            ),
            'YYYY-MM-DD'
          ) AS day,
          COUNT(*) AS sent
        FROM public.attempts a
        WHERE a.survey_id = :sid
          AND a.estado = 'enviado'
          AND (
            :date_from IS NULL OR
            (COALESCE(a.actualizado_en, a.creado_en))::date >= CAST(:date_from AS date)
          )
          AND (
            :date_to IS NULL OR
            (COALESCE(a.actualizado_en, a.creado_en))::date <= CAST(:date_to AS date)
          )
        GROUP BY 1
        ORDER BY 1
    """), {
        "sid": str(survey_id),
        "date_from": date_from,
        "date_to": date_to,
    }).mappings().all()

    return ProgressDailyOut(
        series=[ProgressDay(day=r["day"], sent=int(r["sent"])) for r in rows]
    )

@router.get("/sections/summary", response_model=list[SectionSummaryRow])
def sections_summary(
    survey_id: UUID = Query(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Resumen por sección:
    - n_preguntas (likert en la encuesta)
    - n_respuestas (total de respuestas likert capturadas en la sección)
    - promedio (AVG global de la sección)
    - mejor/peor pregunta de la sección por promedio
    """
    _ensure_survey(db, survey_id)

    rows = db.execute(text("""
        WITH likert AS (
          SELECT q.section_id,
                 q.id  AS question_id,
                 q.codigo,
                 q.enunciado,
                 r.valor_likert::numeric AS val
          FROM public.responses r
          JOIN public.attempts  a ON a.id = r.attempt_id
          JOIN public.questions q ON q.id = r.question_id
          WHERE a.survey_id = :sid
            AND a.estado   = 'enviado'
            AND r.valor_likert IS NOT NULL
            AND q.tipo = 'likert'
        ),
        -- promedio por pregunta dentro de su sección
        perq AS (
          SELECT section_id, question_id, AVG(val) AS avg_q
          FROM likert
          GROUP BY section_id, question_id
        ),
        best AS (
          SELECT section_id, question_id, avg_q,
                 ROW_NUMBER() OVER (PARTITION BY section_id ORDER BY avg_q DESC NULLS LAST) AS rn
          FROM perq
        ),
        worst AS (
          SELECT section_id, question_id, avg_q,
                 ROW_NUMBER() OVER (PARTITION BY section_id ORDER BY avg_q ASC NULLS LAST) AS rn
          FROM perq
        ),
        sec_agg AS (
          SELECT
            s.id     AS section_id,
            s.titulo AS titulo,
            COUNT(DISTINCT q.id)                                   AS n_preguntas,
            COUNT(l.val)                                           AS n_respuestas,
            AVG(l.val)                                             AS promedio
          FROM public.survey_sections s
          JOIN public.questions q ON q.section_id = s.id
          LEFT JOIN likert l      ON l.question_id = q.id
          WHERE s.survey_id = :sid
            AND q.tipo = 'likert'
          GROUP BY s.id, s.titulo
        )
        SELECT
          sa.section_id,
          sa.titulo,
          sa.n_preguntas,
          sa.n_respuestas,
          sa.promedio,

          qbest.id        AS mejor_question_id,
          qbest.codigo    AS mejor_codigo,
          qbest.enunciado AS mejor_enunciado,
          b.avg_q         AS mejor_promedio,

          qworst.id        AS peor_question_id,
          qworst.codigo    AS peor_codigo,
          qworst.enunciado AS peor_enunciado,
          w.avg_q          AS peor_promedio

        FROM sec_agg sa
        LEFT JOIN best  b     ON b.section_id = sa.section_id AND b.rn = 1
        LEFT JOIN public.questions qbest ON qbest.id = b.question_id

        LEFT JOIN worst w     ON w.section_id = sa.section_id AND w.rn = 1
        LEFT JOIN public.questions qworst ON qworst.id = w.question_id

        ORDER BY sa.titulo
    """), {"sid": str(survey_id)}).mappings().all()

    out = []
    for r in rows:
        out.append(SectionSummaryRow(
            section_id = r["section_id"],
            titulo = r["titulo"],
            n_preguntas = int(r["n_preguntas"] or 0),
            n_respuestas = int(r["n_respuestas"] or 0),
            promedio = float(r["promedio"]) if r["promedio"] is not None else None,

            mejor_question_id = r.get("mejor_question_id"),
            mejor_codigo = r.get("mejor_codigo"),
            mejor_enunciado = r.get("mejor_enunciado"),
            mejor_promedio = float(r["mejor_promedio"]) if r.get("mejor_promedio") is not None else None,

            peor_question_id = r.get("peor_question_id"),
            peor_codigo = r.get("peor_codigo"),
            peor_enunciado = r.get("peor_enunciado"),
            peor_promedio = float(r["peor_promedio"]) if r.get("peor_promedio") is not None else None,
        ))
    return out

@router.get("/exports/survey/{survey_id}/responses.csv")
def export_responses_csv(
    survey_id: UUID = Path(..., description="ID de encuesta"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Exporta todas las respuestas de la encuesta en CSV (crudo).
    Columnas:
      attempt_id, user_id, teacher_id, question_id, codigo, section,
      valor_likert, texto_json, enviado_en
    """
    _ensure_survey(db, survey_id)

    rows = db.execute(text("""
      SELECT
        a.id AS attempt_id,
        a.user_id,
        a.teacher_id,
        r.question_id,
        q.codigo,
        s.titulo AS section,
        r.valor_likert,
        r.texto,
        COALESCE(a.actualizado_en, a.creado_en) AS enviado_en
      FROM public.attempts a
      JOIN public.responses r ON r.attempt_id = a.id
      JOIN public.questions q ON q.id = r.question_id
      JOIN public.survey_sections s ON s.id = q.section_id
      WHERE a.survey_id = :sid
        AND a.estado = 'enviado'
      ORDER BY COALESCE(a.actualizado_en, a.creado_en) DESC, a.id, q.orden
    """), {"sid": str(survey_id)}).mappings().all()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "attempt_id","user_id","teacher_id","question_id",
        "codigo","section","valor_likert","texto_json","enviado_en"
    ])

    for r in rows:
        texto_obj = r.get("texto")
        # Aseguramos serialización legible del JSONB
        if texto_obj is None:
            texto_str = ""
        else:
            try:
                texto_str = json.dumps(texto_obj, ensure_ascii=False)
            except Exception:
                texto_str = str(texto_obj)

        w.writerow([
            r["attempt_id"],
            r.get("user_id"),
            r.get("teacher_id"),
            r["question_id"],
            r.get("codigo"),
            r.get("section"),
            r.get("valor_likert"),
            texto_str,
            r.get("enviado_en"),
        ])

    buf.seek(0)
    filename = f"survey-{survey_id}-responses.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/exports/survey/{survey_id}/responses-pretty.csv")
def export_csv_pretty(
    survey_id: UUID = Path(...),
    tz: str = Query("America/Bogota"),
    include_ids: bool = Query(False),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    rows = db.execute(
        text("""
            SELECT
              a.id AS attempt_id,
              a.user_id,
              a.teacher_id,
              r.question_id,

              u.email         AS usuario_email,
              COALESCE(u.nombre, u.email) AS usuario_nombre,

              t.identificador AS docente_identificador,
              t.nombre        AS docente_nombre,
              t.programa      AS docente_programa,

              q.codigo        AS pregunta_codigo,
              q.enunciado     AS pregunta_enunciado,
              s.titulo        AS seccion,

              r.valor_likert  AS valor_likert,
              CASE r.valor_likert
                WHEN 1 THEN 'Muy en desacuerdo'
                WHEN 2 THEN 'En desacuerdo'
                WHEN 3 THEN 'Ni de acuerdo ni en desacuerdo'
                WHEN 4 THEN 'De acuerdo'
                WHEN 5 THEN 'Muy de acuerdo'
                ELSE NULL
              END AS valor_texto,

              r.texto->>'positivos'   AS q16_positivos,
              r.texto->>'mejorar'     AS q16_mejorar,
              r.texto->>'comentarios' AS q16_comentarios,

              to_char(
                (COALESCE(a.actualizado_en, a.creado_en)) AT TIME ZONE :tz,
                'YYYY-MM-DD HH24:MI:SS'
              ) AS enviado_local

            FROM public.attempts a
            JOIN public.responses r ON r.attempt_id = a.id
            JOIN public.questions q ON q.id = r.question_id
            JOIN public.survey_sections s ON s.id = q.section_id
            JOIN public.teachers t ON t.id = a.teacher_id
            LEFT JOIN public.users u ON u.id = a.user_id
            WHERE a.survey_id = :sid
              AND a.estado = 'enviado'
            ORDER BY enviado_local DESC, docente_nombre, pregunta_codigo

        """),
        {"sid": str(survey_id), "tz": tz}
    ).mappings().all()

    import io, csv
    buf = io.StringIO()
    buf.write("\ufeff")  # BOM para Excel

    human_headers = [
        "enviado_local",
        "usuario_email", "usuario_nombre",
        "docente_identificador", "docente_nombre", "docente_programa",
        "pregunta_codigo", "pregunta_enunciado", "seccion",
        "valor_likert", "valor_texto",
        "q16_positivos", "q16_mejorar", "q16_comentarios",
    ]
    id_headers = ["attempt_id", "user_id", "teacher_id", "question_id"] if include_ids else []

    w = csv.writer(buf)
    w.writerow(human_headers + id_headers)

    for r in rows:
        row = [
            r.get("enviado_local") or "",
            r.get("usuario_email") or "",
            r.get("usuario_nombre") or "",
            r.get("docente_identificador") or "",
            r.get("docente_nombre") or "",
            r.get("docente_programa") or "",
            r.get("pregunta_codigo") or "",
            r.get("pregunta_enunciado") or "",
            r.get("seccion") or "",
            r.get("valor_likert") if r.get("valor_likert") is not None else "",
            r.get("valor_texto") or "",
            r.get("q16_positivos") or "",
            r.get("q16_mejorar") or "",
            r.get("q16_comentarios") or "",
        ]
        if include_ids:
            row += [
                r.get("attempt_id"),
                r.get("user_id"),
                r.get("teacher_id"),
                r.get("question_id"),
            ]
        w.writerow(row)

    buf.seek(0)
    from fastapi.responses import StreamingResponse
    filename = f"survey-{survey_id}-responses-pretty.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/exports/survey/{survey_id}/questions.csv")
def export_questions_csv(
    survey_id: UUID = Path(..., description="ID de encuesta"),
    include_stats: bool = Query(True, description="Incluir estadísticas (promedio, desviación, total respuestas)"),
    db: Session = Depends(get_db),
    _admin = Depends(require_admin),
):
    """
    Exporta todas las preguntas de una encuesta con sus configuraciones y opcionalmente estadísticas.
    
    Columnas básicas:
      - question_id, codigo, enunciado, tipo, peso, orden, seccion
    
    Columnas estadísticas (si include_stats=true):
      - total_respuestas, promedio, desviacion_estandar
    """
    _ensure_survey(db, survey_id)
    
    # Query base con información de preguntas
    base_query = """
        SELECT
          q.id AS question_id,
          q.codigo,
          q.enunciado,
          q.tipo,
          q.peso,
          q.orden,
          s.titulo AS seccion,
          s.orden AS seccion_orden
    """
    
    # Si se piden estadísticas, agregamos columnas calculadas
    if include_stats:
        base_query += """,
          COUNT(DISTINCT r.id) FILTER (WHERE a.estado = 'enviado') AS total_respuestas,
          ROUND(AVG(r.valor_likert::numeric) FILTER (WHERE a.estado = 'enviado' AND r.valor_likert IS NOT NULL), 2) AS promedio,
          ROUND(STDDEV(r.valor_likert::numeric) FILTER (WHERE a.estado = 'enviado' AND r.valor_likert IS NOT NULL), 2) AS desviacion_estandar
        """
    
    base_query += """
        FROM public.questions q
        JOIN public.survey_sections s ON s.id = q.section_id
    """
    
    if include_stats:
        base_query += """
        LEFT JOIN public.responses r ON r.question_id = q.id
        LEFT JOIN public.attempts a ON a.id = r.attempt_id AND a.survey_id = :sid
        """
    
    base_query += """
        WHERE q.survey_id = :sid
    """
    
    if include_stats:
        base_query += """
        GROUP BY q.id, q.codigo, q.enunciado, q.tipo, q.peso, q.orden, s.titulo, s.orden
        """
    
    base_query += """
        ORDER BY q.orden
    """
    
    rows = db.execute(text(base_query), {"sid": str(survey_id)}).mappings().all()
    
    # Generar CSV
    buf = io.StringIO()
    buf.write("\ufeff")  # BOM para Excel
    w = csv.writer(buf)
    
    # Headers
    headers = [
        "question_id",
        "codigo",
        "enunciado",
        "tipo",
        "peso",
        "orden",
        "seccion",
    ]
    
    if include_stats:
        headers.extend([
            "total_respuestas",
            "promedio",
            "desviacion_estandar",
        ])
    
    w.writerow(headers)
    
    # Filas
    for r in rows:
        row = [
            str(r["question_id"]),
            r.get("codigo") or "",
            r.get("enunciado") or "",
            r.get("tipo") or "",
            r.get("peso") if r.get("peso") is not None else "",
            r.get("orden") if r.get("orden") is not None else "",
            r.get("seccion") or "",
        ]
        
        if include_stats:
            row.extend([
                r.get("total_respuestas") or 0,
                r.get("promedio") if r.get("promedio") is not None else "",
                r.get("desviacion_estandar") if r.get("desviacion_estandar") is not None else "",
            ])
        
        w.writerow(row)
    
    buf.seek(0)
    filename = f"survey-{survey_id}-questions.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
